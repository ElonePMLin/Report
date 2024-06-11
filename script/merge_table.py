from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment, Protection
import pandas as pd
import chardet
import os


class MergeData:

    HANDLE_COL = {
        "日期": "B", "销售总额": "C", "退单金额": "G", "退单数": "H", "上新数": "J",
        "访客": "K", "下单人数": "L", "下单总件数": "M"
    }

    def __init__(self):
        self.report_data = None
        self.sheet_name = None

    @classmethod
    def read_excel(cls, path, sheet_name=None):
        try:
            suffix = path.split(".")[-1]
            if suffix == "csv":
                encoded = "utf-8"
                with open(path, 'rb') as f:
                    encode = chardet.detect(f.read())
                    if "UTF-8" not in encode["encoding"]:
                        encoded = "GBK"
                    else:
                        encoded = "utf-8"
                return pd.read_csv(path, encoding=encoded)
            else:
                if sheet_name:
                    return pd.read_excel(path, sheet_name=sheet_name)
                else:
                    return pd.read_excel(path)
        except Exception as e:
            print("文件", e)
            raise e

    def merge3file(self, overview_file, report_file, spu_file, store_name, refund_file=None):
        spu = self.spu(spu_file)
        overview = self.overview(overview_file)
        report = self.report(report_file, store_name, overview)  # 以report为主
        spu["日期"] = spu["日期"].astype("datetime64[ns]")
        overview["日期"] = overview["日期"].astype("datetime64[ns]")
        data = overview.merge(spu, how="left").fillna(0)
        if refund_file is not None:
            data = self.add_refund_data(data, refund_file)
            report["退单金额"] = data["退单金额"]
            report["退单数"] = data["退单数"]
        data["上新数"] = data["上新数"].astype("int64")
        report["日期"] = data["日期"]
        report["销售总额"] = data["销售总额"]
        report["访客"] = data["访客"]
        report["下单人数"] = data["下单人数"]
        report["下单总件数"] = data["下单总件数"]
        report["上新数"] = data["上新数"]
        SaveExcel.tmpSave(report_file, self.sheet_name, store_name, report, report_file)
        return report

    def merge2file(self, overview_file, report_file, store_name, refund_file=None):
        overview = self.overview(overview_file)
        if refund_file is not None:
            overview = self.add_refund_data(overview, refund_file)
        report = self.report(report_file, store_name, overview)  # 以report为主
        report["日期"] = overview["日期"]
        report["销售总额"] = overview["销售总额"]
        report["访客"] = overview["访客"]
        report["下单人数"] = overview["下单人数"]
        report["下单总件数"] = overview["下单总件数"]
        # print(report)
        # report = report.fillna(0)
        SaveExcel.tmpSave(report_file, self.sheet_name, store_name, report, report_file)
        return report

    def add_refund_data(self, overview, refund_file):
        refund = self.refund(refund_file)
        refund["日期"] = refund["日期"].astype("datetime64[ns]")
        data = overview.merge(refund, how="left").fillna(0)
        # print(data)
        return data

    def report(self, path, store_name, overview_data):
        if not self.sheet_name:
            return
        try:
            self.read_excel(path, self.sheet_name)
        except Exception as e:
            print("日报表：", e)
            SaveExcel.tmpSave(path, self.sheet_name, store_name, overview_data, path)
        data = self.read_excel(path, self.sheet_name)
        data1 = data.rename(columns=data.iloc[0]).drop([0, 1])[["日期", "销售总额", "上新数", "访客", "下单人数", "下单总件数"]]
        data2 = data.rename(columns=data.iloc[1]).drop([0, 1])[["退单金额", "退单数"]]
        data = data1.merge(data2, left_index=True, right_index=True).iloc[:-1].fillna(0)
        return data.reset_index().drop(columns="index")

    def overview(self, path):
        file_name = path.split("/")[-1].split(".")[0]
        prefix, suffix = file_name.split("年")
        self.sheet_name = ".".join([prefix[-2:], suffix[:3]])
        data = self.read_excel(path)
        data = data.drop(index=0)[['日期', '访客数', '成交客户数', '成交商品件数', '成交金额']].sort_values(by="日期", ascending=True)
        data = data.rename(
            columns={
                "成交金额": "销售总额", "访客数": "访客", "成交客户数": "下单人数", "成交商品件数": "下单总件数"
            }
        )
        # data["日期"] = data["日期"].str.replace("-", "/")
        # data.info()
        data["日期"] = data["日期"].astype("datetime64[ns]")
        data = data.reset_index().drop(columns="index")
        return data

    def spu(self, path):
        data = self.read_excel(path)
        data = data["上架时间"].dt.date.value_counts().reset_index(name="上新数").rename(columns={"上架时间": "日期"})
        return data

    def refund(self, path):
        data = self.read_excel(path)
        data = data[["售后审核时间", "退款金额"]]
        data["日期"] = data["售后审核时间"].astype("datetime64[ns]").dt.date
        data["售后审核时间"] = data["日期"]
        data = data.groupby(by="日期", as_index=False).agg(
            {"退款金额": "sum", "售后审核时间": "count"}
        ).rename(columns={"售后审核时间": "退单数", "退款金额": "退单金额"})
        return data


class SaveExcel:

    def __init__(self):
        pass

    @classmethod
    def tmpSave(cls, path, sheet_name, store_name, data, toPath):
        workbook = load_workbook(path)
        try:
            sheet = workbook[sheet_name]
        except Exception as e:
            print("暂存", e)
            cls.styles(workbook, sheet_name, len(data), path)
            return
        try:
            cells = MergeData.HANDLE_COL
            sheet["A1"] = f"/{sheet_name.split('.')[-1][:-1]}/月份  日报表"
            sheet["A2"] = f"{store_name}旗舰店"
            print(f"/{sheet_name.split('.')[-1][:-1]}/月份  日报表")
            HANDLE_COL = {
                "日期": "B", "销售总额": "C", "退单金额": "G", "退单数": "H",
                "上新数": "J", "访客": "K", "下单人数": "L", "下单总件数": "M"
            }
            data["日期"] = data["日期"].astype("datetime64[ns]")
            # print("123", data)
            for n, date in enumerate(data.index):
                num = 4 + n
                sheet[cells['日期'] + str(num)] = data.loc[date, '日期']
                sheet[cells['日期'] + str(num)].number_format = 'm"月"d"日"'
                sheet[cells['销售总额'] + str(num)] = data.loc[date, '销售总额'].round(1)
                sheet[cells['退单金额'] + str(num)] = data.loc[date, '退单金额'].round(1)
                sheet[cells['退单数'] + str(num)] = int(data.loc[date, '退单数'])
                sheet[cells['上新数'] + str(num)] = int(data.loc[date, '上新数'])
                sheet[cells['访客'] + str(num)] = int(data.loc[date, '访客'])
                sheet[cells['下单人数'] + str(num)] = int(data.loc[date, '下单人数'])
                sheet[cells['下单总件数'] + str(num)] = int(data.loc[date, '下单总件数'])
            workbook.save(toPath)
            return True
        except Exception as e:
            print("暂存2", e)
            return False

    @staticmethod
    def styles(workbook, new_sheet, monthOfday, path):
        form = os.sep.join(os.path.abspath(__file__).split(os.sep)[:-2])
        form += "/template/format.xlsx"
        wb = load_workbook(form)
        # wb = load_workbook(f'/resource/format.xlsx')  # for app
        fm = wb['2023年']
        report = workbook
        report.create_sheet(new_sheet)
        ns = report[new_sheet]
        for idx, row in enumerate(fm.rows):
            if monthOfday + 3 <= idx < 34:
                continue
            for col in row:
                # 值
                if idx == 34 and col.value:
                    ns[f"{col.coordinate[0]}{monthOfday + 4}"].value = col.value.replace("34", str(monthOfday + 3))
                elif idx == 35 and col.value and monthOfday != 31:
                    ns[f"{col.coordinate[0]}{monthOfday + 5}"].value = col.value.replace("35", str(monthOfday + 4))
                elif not str(ns[col.coordinate]).startswith('<MergedCell'):
                    ns[col.coordinate].value = col.value
                # 字体
                font = Font(
                    name=col.font.name,
                    size=col.font.sz,
                    bold=col.font.b,
                    italic=col.font.i,
                    vertAlign=col.font.vertAlign,
                    underline=col.font.u,
                    strike=col.font.strike,
                    color=col.font.color
                )
                ns[col.coordinate].font = font
                # 样式
                fill = PatternFill(
                    fill_type=col.fill.patternType,
                    start_color=col.fill.fgColor,
                    end_color=col.fill.bgColor
                )
                ns[col.coordinate].fill = fill
                # 边界
                border = Border(
                    left=col.border.left,
                    right=col.border.right,
                    top=col.border.top,
                    bottom=col.border.bottom,
                    diagonal=col.border.diagonal,
                    diagonal_direction=col.border.diagonal_direction,
                    outline=col.border.outline,
                    vertical=col.border.vertical,
                    horizontal=col.border.horizontal
                )
                ns[col.coordinate].border = border
                # 位置
                alignment = Alignment(
                    horizontal=col.alignment.horizontal,
                    vertical=col.alignment.vertical,
                    text_rotation=col.alignment.text_rotation,
                    wrap_text=col.alignment.wrap_text,
                    shrink_to_fit=col.alignment.shrink_to_fit,
                    indent=col.alignment.indent
                )
                ns[col.coordinate].alignment = alignment
                # 数值格式
                ns[col.coordinate].number_format = col.number_format
                # 保护样式
                protection = Protection(
                    locked=col.protection.locked,
                    hidden=col.protection.hidden
                )
                ns[col.coordinate].protection = protection
            # ns.merge_cells(merge)
        ns.merge_cells("A1:R1")  # title
        ns.merge_cells(f"A2:A{monthOfday + 3}")  # store(days + 3)
        ns.merge_cells("B2:B3")  # date
        ns.merge_cells("C2:C3")  # sale
        ns.merge_cells("D2:F2")  # num
        ns.merge_cells("G2:H2")  # num
        for col in ["I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "T", "U", "V", "W"]:
            ns.merge_cells(f"{col}2:{col}3")  # total
        workbook.save(path)
        print("创建成功")
        return ns


if __name__ == '__main__':
    merge = MergeData()
    report_path = "../res/屿笙栀/屿笙栀日报表1.xlsx"
    overview_path = "../res/屿笙栀/交易概况_数据概览_全部渠道_2024年01月.xls"
    spu_path = "../res/屿笙栀/导出SPU_2024_01_27_16_51_35（屿笙栀）.xlsx"
    refund_path = "../res/退单（屿笙栀）.xlsx"
    # merge.overview(overview_path)
    # merge.report(report_path)
    # merge.spu(spu_path)
    # merge.refund(refund_path)
    print(merge.merge2file(overview_path, report_path, "卡维妲", refund_path))
    # print(merge.merge3file(overview_path, report_path, spu_path, "卡维妲"))
    # merge.overview(overview_path)
